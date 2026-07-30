"""
Report generation in CSV, XLSX and PDF for any module, plus summary reports.
"""
import io
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse

from app.auth import get_current_user
from app.modules import MODULES, module_headers
from app.storage import get_storage

router = APIRouter(prefix="/reports", tags=["reports"])


def _rows_for(scope: str) -> tuple[list[str], list[dict], str]:
    if scope in MODULES:
        return module_headers(scope), get_storage().list_rows(scope), MODULES[scope]["label"]
    raise HTTPException(status_code=404, detail=f"Unknown report scope '{scope}'")


def _csv(headers, rows) -> bytes:
    import csv
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
    w.writeheader()
    w.writerows(rows)
    return buf.getvalue().encode("utf-8")


def _xlsx(headers, rows, title) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(12, len(h) + 2)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _pdf(headers, rows, title) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    out = io.BytesIO()
    doc = SimpleDocTemplate(out, pagesize=landscape(A4),
                            leftMargin=10 * mm, rightMargin=10 * mm)
    styles = getSampleStyleSheet()
    elems = [Paragraph(f"{title} Report", styles["Title"]),
             Paragraph(datetime.now().strftime("Generated %Y-%m-%d %H:%M"), styles["Normal"]),
             Spacer(1, 8)]
    # Keep the PDF readable: cap to the first ~10 columns.
    hdr = headers[:10]
    data = [hdr] + [[str(r.get(h, ""))[:22] for h in hdr] for r in rows[:200]]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EFF6FF")]),
    ]))
    elems.append(table)
    doc.build(elems)
    return out.getvalue()


@router.get("/{scope}")
def generate(scope: str, format: str = "csv", user: dict = Depends(get_current_user)):
    headers, rows, title = _rows_for(scope)
    fmt = format.lower()
    stamp = datetime.now().strftime("%Y%m%d")
    if fmt == "csv":
        return StreamingResponse(
            io.BytesIO(_csv(headers, rows)), media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{scope}_{stamp}.csv"'})
    if fmt in ("xlsx", "excel"):
        return StreamingResponse(
            io.BytesIO(_xlsx(headers, rows, title)),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{scope}_{stamp}.xlsx"'})
    if fmt == "pdf":
        return StreamingResponse(
            io.BytesIO(_pdf(headers, rows, title)), media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{scope}_{stamp}.pdf"'})
    raise HTTPException(status_code=400, detail="format must be csv, xlsx or pdf")
